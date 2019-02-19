"""
.. module:: gcommands
   :platform: Linux
   :synopsis: Commands Module for GlobalConsole

.. moduleauthor:: Grzegorz Cylny

"""
import os
import re
from multiprocessing.pool import ThreadPool
from pprint import pprint
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import paramiko
from scp import SCPClient
from socket import timeout as SocketTimeout
from time import sleep
#
from globalconsole.gutils import GcUtils as gutils


class GcCommand:
    """This class features:

    -handle GC commands


    """
    def __init__(self, config, logger, vars, creds, hosts):
        """
        Construct object

        Args:
            config (object): holds instance of GcConfig class
            logger (object): holds instance of GcLogging class
            vars (object): holds instance of GcVars class
            creds (object): holds instance of GcCreds class
            hosts (object): holds instance of GcHosts class


        """
        #initialize related objects
        self.gLogging = logger
        self.gConfig = config.config
        self.gVars = vars
        # if hostfile, credfile not provided, use default one from config file
        # start TinyDB object
        self.gHosts = hosts
        self.gCreds = creds
        #list of active connections in a form of list of tuples (hostname, conn_handler)
        self.connections = []
        #list of outputs from a given command, one entry per execution
        self.result = []
        # dictionary mapping ip to hostname, beeing used by scp
        self.scphostdict = {}
        # dictionary mapping hostname to ip and port, beeing used by scp
        self.hostDict = {}
        # configuration of spooling,  being reset after every command run
        self.spool = ""
        # fix to tinydb purging json, caching host and cred info
        self.connhosttempdict = []
        self.conncredtempdict = []
        # fix to tinydb purging json, caching host and cred info
        self.closehosttempdict = []
        #tuple for checking mechanism - stoping batch if element found or not found in results; depened on second tuple elem
        self.check = ([], True)
        #should batch proceed, reset on every step in batch
        self.chain_proceed = 1

    def getConnections(self):
        """
        This method shows active connections. Uses output of pickHosts from GcHosts module

        Examples:

            >>> getConnections()
        """
        self.gLogging.debug("getConnections invoked")
        try:
            if len(self.connections) > 0:
                connected = [x for x, y in self.connections]
                lines = self.gHosts.pickHosts(_printing=False)
                for line in lines:
                    if 'group' in line:
                        #group = gutils.trim_ansi(line).split('id')[0].split(":")[1].strip()
                        group = "id".join(gutils.trim_ansi(line).split('id')[:-1]).split(":")[1].strip()
                    if 'host' in line:
                        #line must be cleaned up from ansi escape sequences
                        host = "id".join(gutils.trim_ansi(line).split('id')[:-1]).split(":")[1].strip()
                        if host in connected:
                            details = self.gHosts.searchHostName(host)[0]
                            print("\t" + host, gutils.color_pick(self.gConfig, '[connected, ip: {}, port: {}]'.format(details['host'], details['port']), self.gConfig['JSON']['pick_yes']))
                        else:
                            print("\t" + host, gutils.color_pick(self.gConfig, '[no connected]', self.gConfig['JSON']['pick_no']))
            else:
                self.gLogging.show("there is no active connection")
        except Exception:
            self.gLogging.error("cannot get connections list")

    def _connectOne(self, hostname):
        """
        This method connects to a given host

        Args:
            hostname (str): hostname to connect

        Returns:
            tuple: hostname and ssh connection handler

        """
        self.gLogging.debug("_connectOne invoked")

        #host = self.gHosts.searchHostName(hostname)[0]
        #cred = self.gCreds.searchCredName(host['def_cred'])[0]
        # fix to tinydb purging json
        try:
            host = [x for x in self.connhosttempdict if x['hostname'] == hostname][0]
        except Exception:
            self.gLogging.error("cannot found given hostname: " + hostname)
            return (host['hostname'], None)

        try:
            cred = [x for x in self.conncredtempdict if x['credname'] == host['def_cred']][0]
        except Exception:
            self.gLogging.error("cannot found valid credentials for: " + hostname)
            return (host['hostname'], None)

        self.gLogging.debug("_connectOne host value: %s, host creds: %s, cred user: %s " % (host['host'], host['def_cred'], cred['username']))
        try:
            if cred['encrypted'] == 'yes':
                passwds = self.gCreds.decryptPasswds(cred)
                passwd = passwds[0]
                key_passwd = passwds[1]
            else:
                passwd = cred['password']
                key_passwd = cred['key_password']
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            if cred['use'] == 'password':
                client.connect(host['host'], port=host['port'], username=cred['username'], password=passwd, timeout=int(self.gConfig['COMMAND']['ssh_timeout']))
                self.gLogging.info("successfully connected to: %s using password" % host['host'])
                stdin, stdout, stderr = client.exec_command(self.gConfig['COMMAND']['hello_command'])
                stdin.close()
                for line in stdout.read().splitlines():
                    self.gLogging.show('%s (%s:%s) says: %s ' % (host['hostname'], host['host'], host['port'], line.decode("utf-8")))
                client.get_transport().set_keepalive(int(self.gConfig['COMMAND']['keep_alive_interval']))
                return (host['hostname'], client)
            elif cred['use'] == 'key':
                privkey = paramiko.RSAKey.from_private_key_file(cred["key"], password=key_passwd)
                client.connect(host['host'], port=host['port'], username=cred['username'], password=passwd, pkey=privkey, timeout=int(self.gConfig['COMMAND']['ssh_timeout']))
                self.gLogging.info("successfully connected to: %s using key" % host['host'])
                stdin, stdout, stderr = client.exec_command(self.gConfig['COMMAND']['hello_command'])
                stdin.close()
                for line in stdout.read().splitlines():
                    self.gLogging.show('%s (%s:%s) says: %s ' % (host['hostname'], host['host'], host['port'], line.decode("utf-8")))
                client.get_transport().set_keepalive(int(self.gConfig['COMMAND']['keep_alive_interval']))
                return (host['hostname'], client)
            else:
                self.gLogging.warning("connection method unknown for host: %s" % host['hostname'])
        except Exception:
            self.gLogging.error("cannot connect to: " + host['hostname'])
            return (host['hostname'], None)

    def _connectOneCallback(self, conn):
        """
        This method adds connection to a pool

        Args:
            conn (tuple): (hostname, shh connection handler)

        """
        self.gLogging.debug("_connectOneCallback invoked")
        try:
            if conn[1] is not None:
                self.connections.append(conn)
        except Exception:
            self.gLogging.error("cannot add connection handler. ")

    def connect(self, hostnames=[]):
        """
        This method creates a connection handler for picked hosts in loaded json file

        Args:
            hostnames (list): list of hostnames to connect to. you can select host even if is not picked-up

        Examples:

            >>> connect()
        """
        self.gLogging.debug("connect invoked")
        try:
            # fix to tinydb purging json
            self.connhosttempdict = self.gHosts.hosttable.all()
            self.conncredtempdict = self.gCreds.credtable.all()
            #
            self.close()
            pool = ThreadPool(processes=int(self.gConfig['COMMAND']['max_threads']))
            if len(hostnames) == 0:
                for i in self.gHosts.hosttable.all():
                    if i['host_checked'] == self.gConfig['JSON']['pick_yes']:
                        pool.apply_async(self._connectOne, args=(i['hostname'],), callback=self._connectOneCallback)
            else:
                for i in hostnames:
                    pool.apply_async(self._connectOne, args=(i,), callback=self._connectOneCallback)
            pool.close()
            pool.join()
        except Exception:
            self.gLogging.error("cannot connect, unhandled error")

    def _commandOne(self, cmd_conn):
        """
        This method executes command on a given host

        Args:
            cmd_conn (tuple): ((hostname, ssh connection handler), command, instance name, database name)

        Returns:
            tuple: (output of command, hostname, command, instance name, database name)

        """
        self.gLogging.debug("_commandOne invoked")
        try:
            self.gLogging.debug("starting thread for host: %s, instance: %s, db: %s" % (cmd_conn[0][0], cmd_conn[2], cmd_conn[3]))
            stdin, stdout, stderr = cmd_conn[0][1].exec_command(cmd_conn[1], get_pty=True, timeout=int(self.gConfig['COMMAND']['ssh_cmd_timeout']))
            stdin.close()
            self.gLogging.debug("stopping thread for host: %s, instance: %s, db: %s" % (cmd_conn[0][0], cmd_conn[2], cmd_conn[3]))
            return (stdout.read(), cmd_conn[0][0], cmd_conn[1], cmd_conn[2], cmd_conn[3], cmd_conn[4])
        except SocketTimeout:
            return ("_GC: TIMEOUT OCCURED_", cmd_conn[0][0], cmd_conn[1], cmd_conn[2], cmd_conn[3], cmd_conn[4])
        except IOError:
            return ("_GC: TIMEOUT OCCURED_", cmd_conn[0][0], cmd_conn[1], cmd_conn[2], cmd_conn[3], cmd_conn[4])
        except Exception:
            print(type(Exception).__name__)
            self.gLogging.error("cannot run command %s at host: %s" % (cmd_conn[1], cmd_conn[0][0]))

    def _commandOneCallback(self, result):
        """
        This method adds result of command to a pool. If ``wait_for_all_hosts`` is set to *NO* it prints result of a command, also.

        Args:
            result (tuple): (output of command, hostname, command, instance name, database name)

        """
        self.gLogging.debug("_commandOneCallback invoked")
        try:
            self.result.append(result)

            if self.gConfig['COMMAND']['wait_for_all_hosts'] == 'NO':
                self._result_show(result)

        except Exception:
            self.gLogging.error("unhandled error in _commandOneCallback")

    def command(self, command, host="all", db2=False, **kwargs):
        """
        This method executes command on a picked hosts in loaded json file.

        Should not be used directly, use ``os`` or ``db2`` instead

        Args:
            command (str): command to execute
            host (str): default is all picked hosts, can be set to one particular host
            db2 (bool): run commands specific for db2
            kwargs (dict): db2 specific variables; user - username to use, env - os variable to set, shell - change default one, placeholder - place db or instance name  in command, level - run command at db or instance level, osmode - should command be run by db2 engine or os

        Examples:

            >>> command("df -h")
        """
        self.gLogging.debug("command invoked")
        self.result = []

        # dictionary for mapping hostname to ip and port
        self.hostDict = {}
        for vhost in self.gHosts.hosttable.all():
            self.hostDict[vhost["hostname"]] = "({}:{})".format(vhost["host"], vhost["port"])

        pool = ThreadPool(processes=int(self.gConfig['COMMAND']['max_threads']))
        if host == "all":
            connections = self.connections
        else:
            connections = [x for x in self.connections if x[0] == host]

        pattern = re.compile("\\\\([a-z]|[A-Z])")
        command = pattern.sub(r'\\\1', command)

        if db2 is True:
            command = command.replace("*", "\*")
            if kwargs['osmode']:
                command = command.replace("\(", "(").replace("\)", ")").replace("\*", "*")

        for conn in connections:
            try:
                if not db2:
                    commands = self.gVars.parseString(command)
                    for cmd_parsed in commands:
                        pool.apply_async(self._commandOne, args=((conn, cmd_parsed, "NA", "NA", "NA"),), callback=self._commandOneCallback)
                        self.gLogging.debug("%s executed command: %s " % (conn[0], cmd_parsed))
                else:
                    vhost = self.gHosts.searchHostName(conn[0])

                    for host in vhost:
                        for install in host["installations"]:
                            for instance in install["instances"]:
                                if kwargs['level'] == 'IN':
                                    if instance['instance_checked'] == self.gConfig['JSON']['pick_yes']:
                                        if kwargs['osmode']:
                                            constr = "; "
                                            terstr = "'"
                                        else:
                                            constr = '; db2 '
                                            terstr = "'"
                                            #command = command.replace("*", "\*")

                                        if len(kwargs['shell']) > 0:
                                            kwargs['shell'] = " -s " + kwargs['shell']

                                        if kwargs['user'] == "current":
                                            cmd = kwargs['env'] + ' . ' + instance["instance_profile"] + constr + command.replace(kwargs['placeholder'], instance['instance_name']).replace("(", "\(").replace(")", "\)") + terstr[:-1]
                                        elif kwargs['user'] == "instance":
                                            cmd = kwargs['env'] + ' sudo su - ' +instance["instance_name"] + kwargs['shell'] + " -c $'" + kwargs['env'] +' . ' +instance["instance_profile"]+ constr + command.replace(kwargs['placeholder'], instance['instance_name']).replace("(", "\(").replace(")", "\)").replace("'", "\\'")+terstr
                                        else:
                                            cmd = kwargs['env'] + ' sudo su - ' + kwargs['user'] + kwargs['shell'] + ' -c "' + kwargs['env'] + ' . ' + instance["instance_profile"] + constr + command.replace(kwargs['placeholder'], instance['instance_name']).replace("(","\(").replace(")", "\)").replace("'", "\\'") + terstr
                                        #print(host['hostname'], cmd)

                                        commands = self.gVars.parseString(cmd)
                                        for cmd_parsed in commands:
                                            #print(cmd_parsed)
                                            pool.apply_async(self._commandOne, args=((conn, cmd_parsed, instance['instance_name'], "NA", install['installation_name']),), callback=self._commandOneCallback)
                                            self.gLogging.debug("%s(%s:%s) executed command: %s " % (host["hostname"], host["host"], host["port"], cmd_parsed))
                                elif kwargs['level'] == 'DB':
                                    for db in instance["databases"]:
                                        if db["db_checked"] == self.gConfig['JSON']['pick_yes']:
                                            if kwargs['osmode']:
                                                constr = "; "
                                                terstr = "'"
                                            else:
                                                constr = '; db2 connect to ' + db["db_name"] + '> /dev/null; db2 '
                                                terstr = " ; db2 terminate > /dev/null'"
                                                #command = command.replace("*", "\*")

                                            if len(kwargs['shell']) > 0:
                                                kwargs['shell'] = " -s " + kwargs['shell']

                                            if kwargs['user'] == "current":
                                                cmd = kwargs['env'] + ' . ' + instance["instance_profile"] + constr + command.replace(kwargs['placeholder'], db['db_name']).replace("(", "\(").replace(")", "\)") + terstr[:-1]
                                            elif kwargs['user'] == "instance":
                                                cmd = kwargs['env'] + ' sudo su - ' +instance["instance_name"] + kwargs['shell'] + " -c $'" + kwargs['env'] +' . ' +instance["instance_profile"]+ constr + command.replace(kwargs['placeholder'], db['db_name']).replace("(", "\(").replace(")", "\)").replace("'", "\\'")+terstr
                                            else:
                                                cmd = kwargs['env'] + ' sudo su - ' + kwargs['user'] + kwargs['shell'] + " -c $'" + kwargs['env'] + ' . ' + instance["instance_profile"] + constr + command.replace(kwargs['placeholder'], db['db_name']).replace("(","\(").replace(")", "\)").replace("'", "\\'") + terstr

                                            #print(host['hostname'], cmd)
                                            commands = self.gVars.parseString(cmd)
                                            for cmd_parsed in commands:
                                                pool.apply_async(self._commandOne, args=((conn, cmd_parsed, instance['instance_name'], db['db_name'], install['installation_name']),), callback=self._commandOneCallback)
                                                self.gLogging.debug("%s(%s:%s) executed command: %s " % (host["hostname"], host["host"], host["port"], cmd_parsed))
                                else:
                                    self.gLogging.info("not supported level")
            except Exception:
                self.gLogging.error("cannot run command, unhandled error ")
        pool.close()
        pool.join()

        self._result_postprocess()

    def set_spool(self, val):
        """
        This method sets spool variable to a given name; sets filename.

        **Supported filename extensions are: xlsx and csv**

        Examples:

            >>> set_spool("result.xlsx")

            >>> set_spool("result.csv")

        """
        self.spool = val

    def _result_show(self, result):
        """
        Internal method to show commands output

        Args:
            result (tuple): result of command execution, (output of command, hostname, command, instance name, database name)

        """
        details = ""
        if result[5] != "NA":
            details += "--install: " + result[5] + "\n"
        if result[3] != "NA":
            details += "--instance: " + result[3]
        if result[4] != "NA":
            details += " --database: " + result[4] + " --"

        self.gLogging.info("--------" + result[1] + " " + self.hostDict[result[1]] + "--------")
        self.gLogging.info(details)
        #self.gLogging.show("")
        for line in result[0].splitlines():
            if len(line.decode("utf-8")) > 0:
                self.gLogging.info(line.decode("utf-8"))
        self.gLogging.show("")

    def _result_postprocess(self):
        """
        Internal method to postprocess result of command


        """
        def _spool_formatter(xresult, option):
            if option == 4:
                return [xresult[1], self.hostDict[xresult[1]], xresult[0], xresult[2]]
            elif option == 5:
                return [xresult[1], self.hostDict[xresult[1]], xresult[5], xresult[3], xresult[0], xresult[2]]
            elif option == 6:
                return [xresult[1], self.hostDict[xresult[1]], xresult[5], xresult[3], xresult[4], xresult[0], xresult[2]]
            else:
                return None

        if self.result[0][3] == "NA" and self.result[0][4] == "NA":
            cols = (["hostname", "host", "result", "command"], 4)
        elif self.result[0][3] != "NA" and self.result[0][4] == "NA":
            cols = (["hostname", "host", "install", "instance", "result", "command"], 5)
        else:
            cols = (["hostname", "host", "install", "instance", "db", "result", "command"], 6)

        #generate excel file
        def xlsx():
            try:
                self.gLogging.debug("generating excel file..")
                wb = Workbook()
                ws = wb.active
                wrap_alignment = Alignment(wrap_text=True, vertical="top")
                byLine = self.gConfig['COMMAND']['spoolxlsxline']

                ws.append(cols[0])
                for xresult in self.result:
                    if byLine == "NO":
                        ws.append(_spool_formatter(xresult, cols[1]))
                    else:
                        for xline in xresult[0].splitlines():
                            if len(xline) > 0:
                                ws.append(_spool_formatter([xresult[1], self.hostDict[xresult[1]], xresult[5], xresult[3], xresult[4], xline.decode("utf-8"), xresult[2]], cols[1]))

                ws.column_dimensions['A'].width = int(self.gConfig['COMMAND']['hostwidth'])
                ws.column_dimensions['B'].width = int(self.gConfig['COMMAND']['hostwidth'])

                if self.result[0][3] == "NA" and self.result[0][4] == "NA":
                    ws.column_dimensions['C'].width = int(self.gConfig['COMMAND']['resultwidth'])
                    ws.column_dimensions['D'].width = int(self.gConfig['COMMAND']['resultwidth'])
                elif self.result[0][3] != "NA" and self.result[0][4] == "NA":
                    ws.column_dimensions['C'].width = int(self.gConfig['COMMAND']['hostwidth'])
                    ws.column_dimensions['D'].width = int(self.gConfig['COMMAND']['hostwidth'])
                    ws.column_dimensions['E'].width = int(self.gConfig['COMMAND']['resultwidth'])
                    ws.column_dimensions['F'].width = int(self.gConfig['COMMAND']['resultwidth'])
                else:
                    ws.column_dimensions['C'].width = int(self.gConfig['COMMAND']['hostwidth'])
                    ws.column_dimensions['D'].width = int(self.gConfig['COMMAND']['hostwidth'])
                    ws.column_dimensions['E'].width = int(self.gConfig['COMMAND']['hostwidth'])
                    ws.column_dimensions['F'].width = int(self.gConfig['COMMAND']['resultwidth'])
                    ws.column_dimensions['G'].width = int(self.gConfig['COMMAND']['resultwidth'])

                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = wrap_alignment

                for c in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']:
                    cell = ws[c]
                    cell.font = Font(bold=True)

                outfile = "{}/{}{}".format(gutils.gcpath(), self.gConfig['COMMAND']['spoolpath'], self.spool)
                wb.save(outfile)
                self.gLogging.info("excel file generated to: %s.. spool turned off" % outfile)
                self.spool = ""
            except Exception:
                self.spool = ""
                self.gLogging.error("cannot generate excel file: %s " % self.gConfig['COMMAND']['spoolpath'] + self.spool)

        #generate csv file
        def plain():
            try:
                self.gLogging.debug("generating spool file..")
                outfile = "{}/{}{}".format(gutils.gcpath(), self.gConfig['COMMAND']['spoolpath'], self.spool)
                with open(outfile, "w") as f:
                    if cols[1] == 4:
                        f.write(self.gConfig['COMMAND']['csv_delimiter'].join(["%s", "%s", "%s", "%s"]) % tuple(cols[0]) + "\n")
                    elif cols[1] == 5:
                        f.write(self.gConfig['COMMAND']['csv_delimiter'].join(["%s", "%s", "%s", "%s", "%s", "%s"]) % tuple(cols[0]) + "\n")
                    else:
                        f.write(self.gConfig['COMMAND']['csv_delimiter'].join(["%s", "%s", "%s", "%s", "%s", "%s", "%s"]) % tuple(cols[0]) + "\n")

                    for xresult in self.result:
                        for xline in xresult[0].splitlines():
                            if len(xline) > 0:
                                if cols[1] == 4:
                                    f.write(self.gConfig['COMMAND']['csv_delimiter'].join(["%s", "%s", "%s", "%s"]) % (xresult[1], self.hostDict[xresult[1]], xline.decode("utf-8"), xresult[2]) + "\n")
                                elif cols[1] == 5:
                                    f.write(self.gConfig['COMMAND']['csv_delimiter'].join(["%s", "%s", "%s", "%s", "%s", "%s"]) % (xresult[1], self.hostDict[xresult[1]], xresult[5], xresult[3], xline.decode("utf-8"), xresult[2]) + "\n")
                                else:
                                    f.write(self.gConfig['COMMAND']['csv_delimiter'].join(["%s", "%s", "%s", "%s", "%s", "%s", "%s"]) % (xresult[1], self.hostDict[xresult[1]], xresult[5], xresult[3], xresult[4], xline.decode("utf-8"), xresult[2]) + "\n")

                self.gLogging.info("csv file generated to: %s.. spool turned off" % outfile)
                self.spool = ""
            except Exception:
                self.spool = ""
                self.gLogging.error("cannot generate spool file: %s " % self.gConfig['COMMAND']['spoolpath'] + self.spool)
        # output is sorted by default by hostname
        # to see this not only in xlsx/csv but in a terminal also one has to set wait_for_all_hosts parameter to YES
        for result in sorted(self.result, key=lambda x: x[1]):
            # self.gLogging.info("--------" + result[1] + " " + self.hostDict[result[1]] + "--------")

            if len(self.check[0]) > 0:
                pattern = re.compile("\s+".join(self.check[0]))
                if self.check[1] is True:
                    if pattern.search(result[0].decode("utf-8")) is None:
                        self.chain_proceed = 0
                else:
                    if pattern.search(result[0].decode("utf-8")) is not None:
                        self.chain_proceed = 0
                self.check = ([], True)

            if self.gConfig['COMMAND']['wait_for_all_hosts'] == 'YES':
                self._result_show(result)
            # self.gLogging.show("")
            # for line in result[0].splitlines():
            #     if len(line.decode("utf-8")) > 0:
            #         self.gLogging.info(line.decode("utf-8"))
            # self.gLogging.show("")

        if os.path.exists("{}/{}".format(gutils.gcpath(), self.gConfig['COMMAND']['spoolpath'])):
            if self.spool.split(".")[-1:][0] == "xlsx":
                xlsx()
            elif self.spool.split(".")[-1:][0] == "csv":
                plain()
            else:
                self.gLogging.debug("no spool or invalid format: xlsx or csv")
        else:
            self.gLogging.info("cannot generate spool file, path do not exists")


    def scan(self, dry=False):
        """
        This method scans connected hosts to get db2 installations, instances and databases.
        Basically is a set of command method invocations.

        Args:
            dry (bool): run without saving changes if True

        Examples:

            >>> scan(dry=False)


        """
        self.gLogging.debug("scan invoked")
        self.gLogging.show("scanning, this may take a few minutes..")
        wait_switched = 0
        if self.gConfig['COMMAND']['wait_for_all_hosts'] == "NO":
            self.gConfig['COMMAND']['wait_for_all_hosts'] = "YES"
            wait_switched = 1
        tmp_profile = ""
        tempinstallations = []

        #installations
        self.result = []
        command = 'sudo su - -c "' + self.gConfig['COMMAND']['db2ls_location'] + '"'
        pool = ThreadPool(processes=int(self.gConfig['COMMAND']['max_threads']))
        input = []
        #todo switch to Command method
        for conn in self.connections:
            try:
                pool.apply_async(self._commandOne, args=((conn, command, "NA", "NA", "NA"),), callback=self._commandOneCallback)
            except Exception:
                self.gLogging.error("cannot run command, unhandled error ")
        pool.close()
        pool.join()

        for install in self.result:
            #print("--------", install[0], "--------")
            for line1 in install[0].splitlines():
                if 'TIMEOUT' not in str(line1):
                    if 'error' not in str(line1):
                        if line1.startswith(bytearray("/", 'utf8')):
                            input.append((line1.decode("utf-8").split(" ")[0] + '/bin/db2ilist', install[1]))
                            tempinstallations.append([install[1], install[2].split('"')[1], line1.decode("utf-8").split(" ")[0]])

        # print("-------install-----------")
        # pprint
        self.gLogging.show("step 1/4 completed..")

        #instances
        self.result = []
        pool = ThreadPool(processes=int(self.gConfig['COMMAND']['max_threads']))

        for conn in self.connections:
            try:
                for x in [x for x, y in input if y == conn[0]]:
                    cmd = 'sudo su - -c "' + x + '"'
                    pool.apply_async(self._commandOne, args=((conn, cmd, "NA", "NA", "NA"),), callback=self._commandOneCallback)
            except Exception:
                self.gLogging.error("cannot run command, unhandled error ")
        pool.close()
        pool.join()

        input = []
        for instance in self.result:
            #print("--------", instance[1], "--------")
            for line2 in instance[0].splitlines():
                if not line2.startswith(bytearray("[", 'utf8')):
                    input.append((line2.decode("utf-8"), instance[1]))
                    tempinstallations.append([instance[1], instance[2].split('"')[1].split("bin")[0][:-1], line2.decode("utf-8")])
        # print("-------instance-----------")
        # pprint(input)
        self.gLogging.show("step 2/4 completed..")

        #profile
        self.result = []
        pool = ThreadPool(processes=int(self.gConfig['COMMAND']['max_threads']))

        for conn in self.connections:
            try:
                for x in [x for x, y in input if y == conn[0]]:
                    cmd = 'sudo su - -c "cat /etc/passwd | grep ^' + x + ': | cut -d: -f6"'
                    pool.apply_async(self._commandOne, args=((conn, cmd, "NA", "NA", "NA"),), callback=self._commandOneCallback)
            except Exception:
                self.gLogging.error("cannot run command, unhandled error ")
        pool.close()
        pool.join()

        input = []

        for profile in self.result:
            #print("--------", profile[1], "--------")
            for line3 in profile[0].splitlines():
                if '[YOU' not in line3.decode("utf-8"):
                    input.append((line3.decode("utf-8") + '/sqllib/db2profile', profile[1]))
                    tempinstallations.append([profile[1], profile[2].split('"')[1].split("|")[1].replace("grep", '').replace("^", '').replace(":", '').strip(), line3.decode("utf-8") + '/sqllib/db2profile'])
        # print("-------profile-----------")
        # pprint(input)
        self.gLogging.show("step 3/4 completed..")

        #database
        self.result = []
        pool = ThreadPool(processes=int(self.gConfig['COMMAND']['max_threads']))

        for conn in self.connections:
            try:
                for x in [x for x, y in input if y == conn[0]]:
                    cmd = 'sudo su - -c ". %s; db2 list db directory"' % x
                    pool.apply_async(self._commandOne, args=((conn, cmd, "NA", "NA", "NA"),), callback=self._commandOneCallback)
            except Exception:
                self.gLogging.error("cannot run command, unhandled error ")
        pool.close()
        pool.join()

        input = []

        for db in self.result:
            #print("--------", db[1], "--------")
            for line4 in re.findall('entry.*?Catalog', str(db[0]), re.DOTALL): #db[0]:
                if "Indirect" in line4:
                    for names in line4.splitlines():
                        if "name" in names:
                            #print(names.split("=")[1].split("Database")[0].split("\\r")[0].strip())
                            input.append((names.split("=")[1].split("Database")[0].split("\\r")[0].strip(), db[1]))
                            tempinstallations.append([db[1], db[2].split('"')[1].split(";")[0].replace(".", '').strip(), names.split("=")[1].split("Database")[0].split("\\r")[0].strip()])
                            cmd = 'sudo su - -c ". %s; db2pd -d %s -hadr"' % (db[2].split('"')[1].split(";")[0].replace(".", '').strip(), names.split("=")[1].split("Database")[0].split("\\r")[0].strip())
                            # pprint([db[1], db[2].split('"')[1].split(";")[0].replace(".", '').strip(), names.split("=")[1].split("Database")[0].split("\\r")[0].strip()])
                            pprint(cmd)
        # print("-------db-----------")
        # pprint(input)
        self.gLogging.show("step 4/4 completed..")

        # print("-------TEMPINSTALLATIONS-----------")
        # pprint(tempinstallations)

        #shape output relaying on above commands' results
        #output is being composed by finding parent-child relations
        try:
            for host, _ in self.connections:
                self.gHosts.updateHost({"scanned": "yes"}, host)

            seen_hosts = set()

            for item in tempinstallations:
                seen_hosts.add(item[0])

            for host in seen_hosts:
                items = []
                for item in tempinstallations:
                    if item[0] == host:
                        items.append([item[1], item[2]])

                l = []
                entries = {}
                #print("--------------", host)
                for fid, id in items:
                    entries[id] = entry = {'id': id, 'fid': fid}
                    if fid == self.gConfig['COMMAND']['db2ls_location']:
                        l.append(entry)
                    else:
                        parent = entries[fid]
                        parent.setdefault('son', []).append(entry)

                tempinstall = list()
                tempinstances = list()
                tempdatabases = list()
                for install in l:
                    for instance in install.get('son', "N"):
                        if instance != "N":
                            for profile in instance.get('son', "N"):
                                if profile != "N":
                                    for db in profile.get('son', "N"):
                                        if db != "N":
                                            tempdatabases.append({'db_name': db['id'], "db_checked": self.gConfig['JSON']['pick_yes'], "db_uuid": gutils.uuid_generator(host+install['id']+instance['id']+db['id'])})
                                        else:
                                            pass
                                    tempinstances.append({'instance_name': instance['id'], "instance_checked": self.gConfig['JSON']['pick_yes'], 'instance_profile': profile['id'], 'databases': tempdatabases, "instance_uuid": gutils.uuid_generator(host+install['id']+instance['id'])})
                                    tempdatabases = []
                                else:
                                    pass
                        else:
                            pass
                    if install != "N":
                        tempinstall.append({'installation_name': install['id'], "installation_checked": self.gConfig['JSON']['pick_yes'], 'instances': tempinstances, "installation_uuid": gutils.uuid_generator(host+install['id'])})
                        tempinstances = []
                if dry:
                    pprint(tempinstall)
                else:
                    self.gHosts.updateHost({"installations": tempinstall, "scanned": "yes"}, host)

            if not dry:
                self.gHosts.pickHosts()

        except Exception:
            self.gLogging.error("cannot update host with scanned data ")

        if wait_switched == 1:
            self.gConfig['COMMAND']['wait_for_all_hosts'] = "NO"

    def _closeOne(self, conn):
        """
        This method closes one active connection

        Args:
            conn (tuple): (hostname, ssh connection handler)

        Returns:
            tuple: returns None if connection successfully removed, tuple (hostname, ssh connection handler) on error

        """
        self.gLogging.debug("_closeOne invoked")

        try:
            #details = self.gHosts.searchHostName(conn[0])[0]
            details = [x for x in self.closehosttempdict if x['hostname'] == conn[0]][0]
            conn[1].close()
            self.connections.remove(conn)
            self.gLogging.info("removed connection to: %s (%s:%s)" % (details['hostname'], details['host'], details['port']))
            return None
        except Exception:
            self.gLogging.error("cannot close connection to host: %s" % conn[0])
            return conn

    def _closeOneCallback(self, conn):
        """
        This method handles _closeOne method output. If conn is not None, add connection back to pool.

        Args:
            conn (tuple): (hostname, ssh connection handler)

        """
        self.gLogging.debug("_closeOneCallback invoked")
        try:
            if conn is not None:
                self.connections.append(conn)
                self.gLogging.info('cannot add connection to %s to connections list ' % conn[0])
        except Exception:
            self.gLogging.error("cannot remove connection to %s from connections list " % conn[0])

    def close(self):
        """
        This method closes active connections

        Examples:

            >>> close()
        """
        self.gLogging.debug("close invoked")

        # fix to tinydb purging json
        self.closehosttempdict = self.gHosts.hosttable.all()

        pool = ThreadPool(processes=int(self.gConfig['COMMAND']['max_threads']))
        for conn in self.connections:
            try:
                pool.apply_async(self._closeOne, args=(conn,), callback=self._closeOneCallback)
            except Exception:
                self.gLogging.error("cannot close, unhandled error ")
        pool.close()
        pool.join()


    def exit(self):
        """
        This method stops program

        Examples:

            >>> exit()
        """
        self.gLogging.debug("exit invoked")
        self.close()
        self.gVars.varsfile.close()
        self.gCreds.credfile.close()
        self.gHosts.hostfile.close()
        self.gLogging.show("bye!")
        self.gLogging.info("Session stopped.")
        exit(0)

    def os(self, cmd, sudo=False, sudoUser="", shell="", repeat=""):
        """
        This method executes os command on a picked hosts in loaded json file

        Args:
            cmd (str): command to execute
            sudo (bool): use sudo if true
            sudoUser (str): choose user for sudo, root is default
            shell (str): choose shell being used
            repeat (str): repeat command execution. <NR_OF_REPETITIONS DELAY_IN_SEC>

        Examples:

            >>> os("ls -la", sudo=False)

            >>> os("cat /etc/passwd", sudo=True, sudoUser="admin", shell="/bin/bash", repeat="10 2")

            using variables:

            >>> os("du -ms {{structs}}", sudo=True)
        """
        self.gLogging.debug("os invoked")

        command = cmd

        if len(shell) > 0:
            shell = " -s " + shell
            command = shell + " -c $'" + cmd.replace("'", "\\'") + "'"
        else:
            shell = ""

        if sudo:
            # same for db2 commands
            command = 'sudo su - ' + sudoUser + shell + " -c $'" + cmd.replace("'", "\\'") + "'"

        try:
            if len(repeat) > 0:
                if gutils.parse_repeat(repeat)[0]:
                    vrepeat = gutils.parse_repeat(repeat)[1]
                    varspool = None
                    for i in range(0, vrepeat[0]):
                        if len(self.spool) > 0:
                            self.spool = gutils.number_to_name(self.spool, i + 1)
                            varspool = self.spool
                        self.command(command)
                        if i < vrepeat[0]-1:
                            if varspool is not None:
                                self.spool = gutils.number_to_name(varspool, i + 2)
                            self.gLogging.info("-----------run:{}--------------\n".format(i+2))
                            sleep(vrepeat[1])
                else:
                    self.gLogging.warning("Invalid format for repeat functionality. Command will run only once")
                    self.command(command)
            else:
                self.command(command)

        except Exception:
            self.gLogging.error("cannot run command: %s" % command)

    #combine features of db2, db2I, db2I2
    def db2(self, command, user="current", env="", shell="", placeholder="{}", level="DB", osmode=False, repeat=""):
        """
        This method executes os command on a picked hosts in loaded json file.

        User can be 'instance' - instance owner, <any> - whoever you want, empty or 'current' - connected user.

        Args:
            command (str): command to run
            user (str): user that should run a command
            env (str): you can prefix command with env variable, etc; example export V=1;
            shell (str): you can choose a shell, example /bin/bash
            placeholder (str): placeholder where database/instance name should be placed
            level (str): choose shell being used
            osmode (bool): run commands at OS level if True (db2 profile loaded)
            repeat (str): repeat command execution. <NR_OF_REPETITIONS DELAY_IN_SEC>

        Examples:

            >>> db2("select * from sysibm.sysdummy1", user="instance")

            >>> db2("get db cfg for {}", user="instance")

            >>> db2("db2pd -db {} -wlocks", user="instance", osmode=True)

            >>> db2("get dbm cfg | grep buffer", user="instance", osmode=False, level='IN')

            >>> db2("echo {}; db2 get dbm cfg | grep buffer", user="instance", osmode=True, level='IN')

            using variables:

            >>> db2("echo {}; db2 get dbm cfg | grep {{structs}}", user="instance", osmode=True, level='IN')

            >>> db2("db2pd -db {} -{{lock}}", user="instance", osmode=True, repeat="5 4")

        """
        self.gLogging.debug("db2 invoked")

        if len(shell) > 0:
            shell = " -s " + shell
        else:
            shell = ""

        if command.endswith(";"):
            command = command[:-1]

        try:
            #self.command(command, db2=True, user=user, env=env, shell=shell, placeholder=placeholder, level=level, osmode=osmode)
            if len(repeat) > 0:
                if gutils.parse_repeat(repeat)[0]:
                    vrepeat = gutils.parse_repeat(repeat)[1]
                    varspool = None
                    for i in range(0, vrepeat[0]):
                        if len(self.spool) > 0:
                            self.spool = gutils.number_to_name(self.spool, i + 1)
                            varspool = self.spool
                        self.command(command, db2=True, user=user, env=env, shell=shell, placeholder=placeholder, level=level, osmode=osmode)
                        if i < vrepeat[0]-1:
                            if varspool is not None:
                                self.spool = gutils.number_to_name(varspool, i + 2)
                            self.gLogging.info("-----------run:{}--------------\n".format(i+2))
                            sleep(vrepeat[1])
                else:
                    self.gLogging.warning("Invalid format for repeat functionality. Command will run only once")
                    self.command(command, db2=True, user=user, env=env, shell=shell, placeholder=placeholder, level=level, osmode=osmode)
            else:
                self.command(command, db2=True, user=user, env=env, shell=shell, placeholder=placeholder, level=level, osmode=osmode)
        except Exception:
            self.gLogging.error("cannot run command: %s" % command)

    def _progress(self, filename, size, sent, peername):
        """
        Internal method to track progress of scp command

        Args:
            filename (str): command to run
            size (int): size of a file in bytes
            sent (int): already sent bytes of a file
            peername (str): ip of host

        """
        try:
            if type(filename) == str:
                self.gLogging.show("object: " + filename + ", size: " + str(size) + ", done: " + str(sent) + ", server: " + self.scphostdict[peername[0]] + ", (" + peername[0] + ":" + str(peername[1]) + ")")
            else:
                self.gLogging.show("object: " + filename.decode("utf-8") + ", size: " + str(size) + ", done: " + str(sent) + ", server: " + self.scphostdict[peername[0]] + ", (" + peername[0] + ":" + str(peername[1]) + ")")
        except Exception:
            self.gLogging.error("cannot track progress")

    def _scpOne(self, cmd_conn):
        """
        This method allows you to scp get/put files/dirs over active connection

        Args:
            cmd_conn (tuple): ((hostname, connection_handler), arg_dictionary)

        where arg_dictionary has: argDict['mode'], argDict['source'], argDict['dest'], argDict['recursive'], argDict['suffix'], argDict['batch']

        """
        self.gLogging.debug("_scpOne invoked")
        try:
            self.gLogging.debug("starting thread for host: %s" % cmd_conn[0][0])
            scp = SCPClient(cmd_conn[0][1].get_transport(), sanitize=lambda x: x, progress4=self._progress)
            if cmd_conn[1]['mode'] == "put":
                self.gLogging.info("scp - command mode: %s, source: %s, dest: %s, recursive: %s, host: %s" % (cmd_conn[1]['mode'], cmd_conn[1]['source'], cmd_conn[1]['dest'], str(cmd_conn[1]['recursive']), cmd_conn[0][0]))
                scp.put(cmd_conn[1]['source'], cmd_conn[1]['dest'], recursive=cmd_conn[1]['recursive'])
            elif cmd_conn[1]['mode'] == "get":
                if cmd_conn[1]['suffix']:
                    #todo check is dest dir
                    cmd_conn[1]['dest'] = cmd_conn[1]['dest'] + os.path.split(cmd_conn[1]['source'])[-1:][0] + "_" + cmd_conn[0][0]
                self.gLogging.info("scp - command mode: %s, source: %s, dest: %s, recursive: %s, host: %s" % (cmd_conn[1]['mode'], cmd_conn[1]['source'], cmd_conn[1]['dest'], str(cmd_conn[1]['recursive']), cmd_conn[0][0]))
                scp.get(cmd_conn[1]['source'], cmd_conn[1]['dest'], recursive=cmd_conn[1]['recursive'])
            else:
                pass
            self.gLogging.debug("stopping thread for host: %s" % cmd_conn[0][0])
        except Exception:
            self.gLogging.error("cannot scp")

    def scp(self, mode, source, dest, recursive=False, suffix=True, batch=False):
        """
        This method allows you to scp get/put files/dirs over active connections

        Args:
            mode (str): valid is 'put' or 'get'
            source (str): source file/dir
            dest (str): destination file/dir
            recursive (bool): should scp look into directory
            suffix (bool): append IP to downloaded dirs/files
            batch (bool): will not ask for confirmation, if set

        Examples:

            >>> scp(mode="get", source="tmp/myfile1.txt", dest="downloads/", recursive=False, suffix=True, batch=False)

            using variables:

            >>> scp(mode="put", source="{{srcs}}", dest="/tmp", recursive=False, suffix=True, batch=True)

        """
        self.gLogging.debug("scp invoked")
        self.gLogging.warning("EXISTING FILES WILL BE OVERWRITTEN, CONTINUE? [Y/n]")
        if batch:
            ans = "Y"
        else:
            ans = input("answer: ")
        if ans == "Y":
            try:
                pool = ThreadPool(processes=int(self.gConfig['COMMAND']['max_threads']))
                self.scphostdict = {}
                for host in self.gHosts.hosttable.all():
                    self.scphostdict[host["host"]] = host["hostname"]
                for conn in self.connections:
                    tempDest = dest
                    tempSrc = source
                    for xdest in self.gVars.parseString(tempDest):
                        for xsrc in self.gVars.parseString(tempSrc):
                            pool.apply_async(self._scpOne, args=((conn, {'mode': mode, 'source': xsrc, 'dest': xdest, 'recursive': recursive, 'suffix': suffix, 'batch': batch}),))
                pool.close()
                pool.join()
            except Exception:
                self.gLogging.error("cannot run scp command")
        else:
            self.gLogging.info("scp aborted. ")


